import openpyxl
import sys
import os


def analyze_results(xlsx_path: str):
    if not os.path.exists(xlsx_path):
        print(f"[FAIL] File not found: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    
    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 60}")
        print(f"  SHEET: {sheet_name}")
        print(f"{'=' * 60}")
        ws = wb[sheet_name]
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("  (empty)")
            continue
        
        headers = list(rows[0])
        print(f"  Columns ({len(headers)}): {', '.join(str(h) for h in headers)}")
        print(f"  Data rows: {len(rows) - 1}")
        
        if len(rows) > 1:
            print(f"\n  First row:")
            for h, v in zip(headers, rows[1]):
                print(f"    {h}: {v}")

    print(f"\n{'=' * 60}")
    print(f"  Summary: {len(wb.sheetnames)} sheets")
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"    {name}: {max(0, len(rows) - 1)} rows")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "evaluation_results.xlsx"
    analyze_results(path)
