"""CluboraX live AI-advisor chat evaluation runner.

What it does:
- Reads questions from `CluboraX_Evaluation_v2.xlsx` (sheet: "Evaluation")
- Calls the live POST /api/ai-advisor/chat/ endpoint for each question
- Records rag_kind + rag_distance from response metadata
- Exports an .xlsx report with: Question, expected tier, source in data,
  actual response, actual tier, rag_kind, rag_distance, pass/fail.

Pass/Fail logic (fixed):
- Tier 1: PASS if rag_kind == "retrieved"
- Tier 2: PASS if rag_kind is "retrieved" OR "generated" (answered something)
          FAIL if refused, errored, or timed out
- Tier 3: PASS if rag_kind == "refused"

Auth:
- The chat endpoint requires authentication (Bearer JWT).
- Provide either:
  - --token (JWT access token), OR
  - --email + --password to fetch a token from /api/auth/login/.

Usage examples:
  python aichatbot/run_live_evaluation.py --base-url http://127.0.0.1:8000 --email you@x.com --password "..."

Environment variables (optional):
- CLUBORAX_BASE_URL
- CLUBORAX_TOKEN
- CLUBORAX_EMAIL
- CLUBORAX_PASSWORD
"""

from __future__ import annotations

import argparse
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import urlsplit
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


DEFAULT_INPUT = Path("aichatbot") / "CluboraX_Evaluation_v2.xlsx"
DEFAULT_SHEET = "Evaluation"
CHAT_PATH = "/api/ai-advisor/chat/"
LOGIN_PATH = "/api/auth/login/"

# Colors for output xlsx
PASS_FILL = PatternFill("solid", fgColor="D9EAD3")
FAIL_FILL = PatternFill("solid", fgColor="FCE5CD")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")


@dataclass(frozen=True)
class EvalRow:
    number: int
    question: str
    expected_tier_raw: str
    source_in_data: str


@dataclass(frozen=True)
class ChatResult:
    actual_response: str
    rag_kind: Optional[str]
    rag_distance: Optional[float]
    error: bool = False


def _normalize_base_url(base_url: str) -> str:
    base_url = (base_url or "").strip()
    if not base_url:
        raise ValueError("base_url is required")
    parts = urlsplit(base_url)
    if not parts.scheme or not parts.netloc:
        raise ValueError(f"base_url must include scheme and host, got: {base_url!r}")
    return f"{parts.scheme}://{parts.netloc}"


def _tier_number(value: Any) -> Optional[int]:
    if value is None:
        return None
    m = re.search(r"\btier\s*(\d)\b", str(value), flags=re.IGNORECASE)
    if not m:
        return None
    n = int(m.group(1))
    return n if n in (1, 2, 3) else None


def _actual_tier_from_rag_kind(rag_kind: Optional[str]) -> Optional[int]:
    if not rag_kind:
        return None
    k = str(rag_kind).strip().lower()
    if k == "retrieved":
        return 1
    if k == "generated":
        return 2
    if k == "refused":
        return 3
    return None


def _actual_tier_label(rag_kind: Optional[str]) -> str:
    k = (rag_kind or "").strip().lower()
    if k == "retrieved":
        return "Tier 1 - Retrieved"
    if k == "generated":
        return "Tier 2 - AI Generated"
    if k == "refused":
        return "Tier 3 - Refused"
    if k == "warming_up":
        return "Warming Up"
    if rag_kind:
        return f"Unknown ({rag_kind})"
    return "Error / Timeout"


def _verdict(expected_tier: Optional[int], rag_kind: Optional[str], error: bool, source_in_data: str = "") -> str:
    """
    Pass/fail logic:
    - Tier 1: PASS only if retrieved
    - Tier 2 (in data):     PASS if retrieved OR generated
    - Tier 2 (not in data): PASS only if generated
                            FAIL if retrieved (wrong answer from weak match)
    - Tier 3: PASS only if refused
    - RETRY if warming_up (excluded from pass rate)
    """
    if error or rag_kind is None:
        return "FAIL"

    k = str(rag_kind).strip().lower()

    if k == "warming_up":
        return "RETRY"

    if expected_tier == 1:
        return "PASS" if k == "retrieved" else "FAIL"

    if expected_tier == 2:
        not_in_data = "not in data" in source_in_data.lower()
        if not_in_data:
            # Must generate — retrieval means wrong answer was returned
            return "PASS" if k == "generated" else "FAIL"
        else:
            # In data — retrieved or generated both acceptable
            return "PASS" if k in ("retrieved", "generated") else "FAIL"

    if expected_tier == 3:
        return "PASS" if k == "refused" else "FAIL"

    return ""


def _find_header_row(ws, required_headers: Iterable[str], search_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    required = {h.strip().lower() for h in required_headers}
    for row_idx in range(1, search_rows + 1):
        row_values = [c.value for c in ws[row_idx]]
        normalized = [("" if v is None else str(v)).strip().lower() for v in row_values]
        header_to_col: Dict[str, int] = {}
        for col_idx, name in enumerate(normalized, start=1):
            if name:
                header_to_col[name] = col_idx
        if required.issubset(set(header_to_col.keys())):
            return row_idx, header_to_col
    raise ValueError(
        f"Could not find header row with: {sorted(required)} (searched first {search_rows} rows)"
    )


def load_eval_rows_xlsx(path: Path, sheet_name: str = DEFAULT_SHEET) -> List[EvalRow]:
    if not path.exists():
        raise FileNotFoundError(str(path))
    wb = load_workbook(path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row, header_to_col = _find_header_row(
        ws,
        required_headers=["question", "expected tier", "source in data"],
        search_rows=30,
    )
    q_col = header_to_col["question"]
    expected_col = header_to_col["expected tier"]
    source_col = header_to_col["source in data"]
    num_col = header_to_col.get("#", None)

    rows: List[EvalRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        q = ws.cell(row=r, column=q_col).value
        question = ("" if q is None else str(q)).strip()
        if not question:
            continue
        num_val = ws.cell(row=r, column=num_col).value if num_col else len(rows) + 1
        try:
            num = int(num_val) if num_val is not None else len(rows) + 1
        except (ValueError, TypeError):
            num = len(rows) + 1

        rows.append(EvalRow(
            number=num,
            question=question,
            expected_tier_raw=("" if ws.cell(row=r, column=expected_col).value is None
                               else str(ws.cell(row=r, column=expected_col).value)).strip(),
            source_in_data=("" if ws.cell(row=r, column=source_col).value is None
                            else str(ws.cell(row=r, column=source_col).value)).strip(),
        ))
    if not rows:
        raise ValueError("No questions found in evaluation sheet")
    return rows


def login_for_token(session: requests.Session, base_url: str, email: str, password: str, timeout_s: int) -> str:
    url = f"{base_url}{LOGIN_PATH}"
    resp = session.post(url, json={"email": email, "password": password}, timeout=timeout_s)
    if resp.status_code != 200:
        raise RuntimeError(f"Login failed ({resp.status_code}): {resp.text}")
    data = resp.json()
    token = (((data or {}).get("data") or {}).get("access_token"))
    if not token:
        raise RuntimeError(f"Login succeeded but no access_token found: {data}")
    return str(token)


def call_chat(
    session: requests.Session,
    base_url: str,
    token: str,
    message: str,
    mode: str,
    session_id: str,
    timeout_s: int,
) -> ChatResult:
    url = f"{base_url}{CHAT_PATH}"
    headers = {"Authorization": f"Bearer {token}"}
    payload: Dict[str, Any] = {"message": message, "mode": mode, "session_id": session_id}

    try:
        resp = session.post(url, json=payload, headers=headers, timeout=timeout_s)
    except requests.RequestException as e:
        return ChatResult(actual_response=f"REQUEST ERROR: {e}", rag_kind=None, rag_distance=None, error=True)

    if resp.status_code != 200:
        text = resp.text[:5000] + ("..." if len(resp.text) > 5000 else "")
        return ChatResult(actual_response=f"HTTP {resp.status_code}: {text}", rag_kind=None, rag_distance=None, error=True)

    data = resp.json() if resp.content else {}
    actual_response = (data.get("message") or "").strip()
    meta = data.get("meta") or {}
    rag_kind = meta.get("rag_kind")
    rag_distance = meta.get("rag_distance")
    try:
        rag_distance_f = float(rag_distance) if rag_distance is not None else None
    except (TypeError, ValueError):
        rag_distance_f = None

    return ChatResult(
        actual_response=actual_response,
        rag_kind=str(rag_kind) if rag_kind is not None else None,
        rag_distance=rag_distance_f,
        error=False,
    )


def export_results_xlsx(
    output_path: Path,
    rows: List[EvalRow],
    results: List[ChatResult],
    threshold_label: str = "",
) -> None:
    if len(rows) != len(results):
        raise ValueError("rows/results length mismatch")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:I1")
    title = f"CluboraX Chatbot Evaluation Results"
    if threshold_label:
        title += f" — Threshold: {threshold_label}"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    headers = ["#", "Question", "Expected Tier", "Source in Data",
               "Actual Response", "Actual Tier", "rag_kind", "rag_distance", "Pass/Fail"]
    ws.append([])  # blank row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 25

    pass_count = 0
    retry_count = 0
    fail_count = 0

    for row, res in zip(rows, results):
        expected_tier_num = _tier_number(row.expected_tier_raw)
        actual_tier_label = _actual_tier_label(res.rag_kind)
        verdict = _verdict(expected_tier_num, res.rag_kind, res.error, row.source_in_data)

        if verdict == "PASS":
            pass_count += 1
            fill = PASS_FILL
        elif verdict == "RETRY":
            retry_count += 1
            fill = WARN_FILL
        else:
            fail_count += 1
            fill = FAIL_FILL

        data_row = [
            row.number,
            row.question,
            row.expected_tier_raw,
            row.source_in_data,
            res.actual_response,
            actual_tier_label,
            res.rag_kind,
            res.rag_distance,
            verdict,
        ]
        ws.append(data_row)
        r = ws.max_row
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 50

    # Column widths
    for col, width in enumerate([4, 45, 22, 16, 55, 20, 14, 14, 12], 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width

    # Summary section below results
    total = len(rows)
    answered = pass_count + retry_count
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)
    summaries = [
        ("Total questions", total),
        ("PASS", pass_count),
        ("FAIL", fail_count),
        ("RETRY (warming up — rerun these)", retry_count),
        ("Pass rate (excl. RETRY)", f"{pass_count}/{total - retry_count} = {pass_count/(total-retry_count)*100:.1f}%" if (total - retry_count) > 0 else "N/A"),
    ]
    for i, (label, val) in enumerate(summaries, 1):
        ws.cell(row=summary_row + i, column=1, value=label)
        ws.cell(row=summary_row + i, column=2, value=str(val))

    # Per-tier summary
    tier_row = summary_row + len(summaries) + 2
    ws.cell(row=tier_row, column=1, value="Per-Tier Summary").font = Font(bold=True)
    tier_headers = ["Tier", "Pass", "Fail", "Retry", "Pass Rate"]
    for col, h in enumerate(tier_headers, 1):
        ws.cell(row=tier_row + 1, column=col, value=h).font = Font(bold=True)

    for tier_num, tier_label in [(1, "Tier 1 — Retrieved"), (2, "Tier 2 — AI Generated"), (3, "Tier 3 — Refused")]:
        tier_rows = [(r, res) for r, res in zip(rows, results)
                     if _tier_number(r.expected_tier_raw) == tier_num]
        t_pass = sum(1 for r, res in tier_rows if _verdict(tier_num, res.rag_kind, res.error, r.source_in_data) == "PASS")
        t_fail = sum(1 for r, res in tier_rows if _verdict(tier_num, res.rag_kind, res.error, r.source_in_data) == "FAIL")
        t_retry = sum(1 for r, res in tier_rows if _verdict(tier_num, res.rag_kind, res.error, r.source_in_data) == "RETRY")
        denom = len(tier_rows) - t_retry
        rate = f"{t_pass}/{denom} = {t_pass/denom*100:.1f}%" if denom > 0 else "N/A"
        r_idx = tier_row + 2 + (tier_num - 1)
        for col, val in enumerate([tier_label, t_pass, t_fail, t_retry, rate], 1):
            ws.cell(row=r_idx, column=col, value=val)

    wb.save(output_path)


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Run CluboraX AI-advisor live chat evaluation.")
    p.add_argument("--base-url", default=os.getenv("CLUBORAX_BASE_URL", "http://127.0.0.1:8000"))
    p.add_argument("--input", default=str(DEFAULT_INPUT))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--token", default=os.getenv("CLUBORAX_TOKEN", ""))
    p.add_argument("--email", default=os.getenv("CLUBORAX_EMAIL", ""))
    p.add_argument("--password", default=os.getenv("CLUBORAX_PASSWORD", ""))
    p.add_argument("--mode", default="general")
    p.add_argument("--session-id", default="eval")
    p.add_argument("--timeout", type=int, default=60, help="Request timeout in seconds (default: 60)")
    p.add_argument("--delay", type=float, default=1.0, help="Seconds to wait between requests (default: 1.0)")
    p.add_argument("--threshold-label", default="", help="Label for this threshold config e.g. '0.6-0.9'")
    default_out = Path("aichatbot") / f"eval_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    p.add_argument("--output", default=str(default_out))
    return p


def main(argv=None) -> int:
    args = build_arg_parser().parse_args(argv)

    try:
        base_url = _normalize_base_url(args.base_url)
    except Exception as e:
        print(f"ERROR: {e}")
        return 2

    input_path = Path(args.input)
    output_path = Path(args.output)
    rows = load_eval_rows_xlsx(input_path, sheet_name=args.sheet)
    print(f"Loaded {len(rows)} questions from {input_path}")
    print(f"Threshold: {args.threshold_label or 'not specified'}")
    print(f"Timeout: {args.timeout}s | Delay between requests: {args.delay}s")
    print("-" * 60)

    with requests.Session() as session:
        token = (args.token or "").strip()
        if not token:
            if args.email and args.password:
                print(f"Logging in as {args.email}...")
                token = login_for_token(session, base_url, args.email, args.password, timeout_s=args.timeout)
                print("Login successful.\n")
            else:
                print("ERROR: No auth. Provide --token OR --email + --password.")
                return 2

        results: List[ChatResult] = []
        for i, row in enumerate(rows, start=1):
            res = call_chat(
                session,
                base_url=base_url,
                token=token,
                message=row.question,
                mode=args.mode,
                session_id=args.session_id,
                timeout_s=args.timeout,
            )
            expected_tier_num = _tier_number(row.expected_tier_raw)
            verdict = _verdict(expected_tier_num, res.rag_kind, res.error, row.source_in_data)
            status_icon = "✅" if verdict == "PASS" else ("⚠️ " if verdict == "RETRY" else "❌")
            print(f"[{i:02d}/{len(rows)}] {status_icon} {verdict:5s} | rag_kind={res.rag_kind} dist={res.rag_distance}")
            results.append(res)
            if args.delay > 0 and i < len(rows):
                time.sleep(args.delay)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    export_results_xlsx(output_path, rows, results, threshold_label=args.threshold_label)

    pass_count = sum(1 for r, res in zip(rows, results)
                     if _verdict(_tier_number(r.expected_tier_raw), res.rag_kind, res.error, r.source_in_data) == "PASS")
    retry_count = sum(1 for r, res in zip(rows, results)
                      if _verdict(_tier_number(r.expected_tier_raw), res.rag_kind, res.error, r.source_in_data) == "RETRY")
    denom = len(rows) - retry_count
    print(f"\n{'='*60}")
    print(f"Results: {pass_count}/{denom} passed ({pass_count/denom*100:.1f}%)" if denom > 0 else "No valid results")
    if retry_count:
        print(f"⚠️  {retry_count} questions hit 'warming_up' — rerun after backend is fully loaded")
    print(f"Saved: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())